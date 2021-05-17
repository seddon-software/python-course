'''
Generate pattern
'''

import numpy as np
import matplotlib.pyplot as plt
import openpyxl, re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import Cell
from PIL import ImageColor
from io import StringIO
import sys

result = StringIO()
old_stdout = sys.stdout  
sys.stdout = result

FILENAME = "data/RugPattern"

def getActiveWorkbookAndSheet(filename):
    wb = load_workbook(filename=f'data/{filename}.xlsx')
    return wb, wb.active

def getCellColor(cell):
    parameters = cell.fill.__repr__()
    parameterList = parameters.split("\n")
    pattern = re.compile(r"^rgb='(.+?)'.*$")
    matcher = pattern.search(parameterList[4])
    result = matcher.group(1)
    return result[2:]       # remove transparency

codes = {
    (101,  67,  33):'d',
    (255, 165,   0):'o',
    (255, 182, 193):'p',
    (152, 251, 152):'l',
    (185, 185, 185):'b',
    (255, 255, 255):'w',
    ( 34, 139,  34):'g',
    (255, 255,   0):'y',
    (  0,   0, 255):'t',
    (165,  42,  42):'m',
    (  0,   0,   0):'x'
}

wb, ws = getActiveWorkbookAndSheet("MASTER")

ROWS = 160*2
COLS = 99*2

rows = (0, ROWS)
cols = (140, COLS)


for row in range(1, ROWS+1):
    print(f"{row:03}", end=" ")

    for col in range(1, COLS+1):
        color = f"#{getCellColor(ws.cell(row, col))}"
        rgb = ImageColor.getcolor(color, "RGB")
        try:
            code = codes[rgb]
            print(code, end="")
        except:
            print(row, col, "unknown color")
        if col%5 == 0: print(" ", end="")
        if col%70 == 0: 
            print(" ")
            print("    ", end="")
    print()
    print()

text = result.getvalue()
sys.stdout = old_stdout

inFile = open(FILENAME, "w")
inFile.write(text)
inFile.close()
print("pattern updated")

