import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import Cell
import openpyxl, re
from PIL import ImageColor

''' done lines 1 to 69 
    next line 70
'''

codes = {
    'd':(101,  67,  33),
    'o':(255, 165,   0),
    'p':(255, 182, 193),
    'l':(152, 251, 152),
    'b':(185, 185, 185),
    'w':(255, 255, 255),
    'g':( 34, 139,  34),
    'y':(255, 255,   0),
    't':(  0,   0, 255),
    'm':(165,  42,  42),
}

def getActiveWorkbookAndSheet(filename):
    wb = load_workbook(filename=f'data/{filename}.xlsx')
    return wb, wb.active

def getCellColor(cell):
    parameters = cell.fill.__repr__()
    parameterList = parameters.split("\n")
    pattern = re.compile(r"^rgb='(.+?)'.*$")
    matcher = pattern.search(parameterList[4])
    result = matcher.group(1)
    return result

def duplicate(d):
    sourceRowStart = ws[d['start']].row
    sourceColStart = ws[d['start']].column
    sourceRowEnd = ws[d['end']].row
    sourceColEnd = ws[d['end']].column
    targetRowStart = ws[d['target']].row
    targetColStart = ws[d['target']].column
    
    rowOffset = targetRowStart - sourceRowStart
    colOffset = targetColStart - sourceColStart

    for row in range(sourceRowStart, sourceRowEnd+1):
        for col in range(sourceColStart, sourceColEnd+1):
            color = getCellColor(ws.cell(row, col))
            fill = PatternFill(start_color=f"{color}", fill_type = "solid")
            ws.cell(row+rowOffset, col+colOffset).fill = fill

def diff(first, second):
    '''
    first is cell name:   "XX88"
    second is cell name:  "YY99"
    '''
    return (ws[first].row    - ws[second].row, 
            ws[first].column - ws[second].column)

def newCell(base, offset):
    '''
    base is cell name:  "XX99"
    offset is a tuple:  (rows, cols)
    '''
    row, column = offset
    newRow = ws[base].row + row
    newColumn = ws[base].column + column
    cell = ws.cell(newRow, newColumn)
    return cell.coordinate

ROWS = 160*2
COLS = 99*2

# quad1   ROWS   1...160    COLS   1... 99
# quad2   ROWS 160...  1    COLS 198...100
# quad3   ROWS 161...320    COLS   1... 99
# quad4   ROWS 161...320    COLS 198...100
#                 321               199
def duplicateQuadrants():
    # quadrant 1
    for row1 in range(1, 161):
        row2 = row1
        for col1 in range(1, 100):
            col2 = 199 - col1
            color = getCellColor(ws.cell(row2, col2))
            fill = PatternFill(start_color=f"{color}", fill_type = "solid")
            ws.cell(row1, col1).fill = fill
    # quadrant 3
    for row3 in range(161, 321):
        row2 = 321 - row3
        for col3 in range(1, 100):
            col2 = 199 - col3
            color = getCellColor(ws.cell(row2, col2))
            fill = PatternFill(start_color=f"{color}", fill_type = "solid")
            ws.cell(row3, col3).fill = fill
    # quadrant 4
    for row4 in range(161, 321):
        row2 = 321 - row4
        for col4 in range(100, 199):
            col2 = col4
            color = getCellColor(ws.cell(row2, col2))
            fill = PatternFill(start_color=f"{color}", fill_type = "solid")
            ws.cell(row4, col4).fill = fill

def writeToCell(cell, color):
    row = ws[cell].row
    col = ws[cell].column
    rgb = codes[color]
    color = f"00{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
    fill = PatternFill(start_color=f"{color}", fill_type = "solid")
    ws.cell(row, col).fill = fill
            
wb, ws = getActiveWorkbookAndSheet("MASTER")
corrections = {
    "DF77" :'y',
    "DF78" :'y',
    "DE79" :'y',
    "DK101":'y',
    "DL101":'y',
    
    
    "CV66" :'t',
    "CW66" :'t',
    "CW132":'b',
    "CX44" :'b',
    "CZ68" :'m',
    "CZ158":'b',
    "CZ159":'b',

    "DA68" :'y',
    "DA69" :'y',
    "DA70" :'y',
    "DB67" :'y',
    "DB69" :'y',
    "DB70" :'y',
    "DC67" :'y',
    "DC68" :'y',
    "DD58" :'d',
    "DD66" :'m',
    "DD67" :'y',
    "DD68" :'y',
    "DE151":'t',
    "DE152":'t',
    "DE153":'t',
    "DE154":'t',
    "DF151":'t',
    "DF152":'t',    
    "DF153":'t',
    "DJ66" :'b',
    "DK66" :'b',
    "DK67" :'b',
    "DK68" :'b',
    "DO84" :'m',
    "DS53" :'b',
    "DY79" :'t',
    "DY80" :'l',

    "EB73" :'t',
    "EB9"  :'y',

    "EB107":'y',
    "EA107":'y',
    "EA108":'y',
    "DZ108":'y',

    "EC73" :'y',
    "EC74" :'y',
    "EC75" :'y',
    "EB74" :'y',
    "ED74" :'y',
    
    "EL74" :'y',
    "EM73" :'y',
    "EM74" :'y',
    "EM75" :'y',
    "EN74" :'y',

    "CW74" : 'y',

    "CX136": 'y',
    "CX137": 'y',
    "CX138": 'y',
    "CY139": 'y',
    "CZ138": 'y',
    "DA138": 'y',
    "DB138": 'y',
    "DB139": 'y',
    "DA140": 'y',
    "CZ141": 'y',
    "CZ142": 'y',
    "CZ143": 'y',
    "DB129": 'b',
    "DA144": 'y',
    "DB145": 'y',
    "DC146": 'y',
    "DD146": 'y',
    "DE145": 'y',
    "DF141": 'y',
    "DG141": 'y',
    "DH141": 'y',
    "DI142": 'y',
    "DJ143": 'y',
    "DK142": 'y',
    "DL142": 'y',
    "DM142": 'y',
    "DN142": 'y',
    "DN143": 'y',
    "DN144": 'y',
    "DN145": 'y',
    "DN147": 'y',
    "DM146": 'y',
    "DL150": 'y',
    "DK151": 'y',
    "DJ152": 'y',
    "DJ153": 'y',
    "DJ154": 'y',

    "DD150": 'y',
    "DE150": 'y',
    "DF150": 'y',
    "DG151": 'y',
    "DG152": 'y',
    "DG153": 'y',
    "DF157": 'y',
    "DF154": 'l',
    "DE155": 'l',


    "EC28" :'m',
    "EF66" :'d',
    "EG66" :'d',
    "EH74" :'y',
    "EH123":'d',
    "EI97" :'g',
    "EJ98" :'g',
    "EN90" :'o',
    "EN91" :'o',
    "EN140":'m',
    "EO5"  :'b',
    "EO90" :'o',
    "EO91" :'o',
    "EO136":'d',

    "EP89" :'g',
    "EP90" :'p',
    "EP133":'g',
    "EP134":'g',
    "EP135":'l',
    "EP136":'d',
    "EQ84" :'d',
    "EQ90" :'g',
    "EQ130":'l',
    "EQ134":'g',
    "EQ135":'g',
    "EQ137":'d',
    "ER69" :'p',
    "ER80" :'d',
    "ER88" :'y',
    "ER91" :'d',
    "ER130":'l',
    "ER135":'g',
    "ER137":'d',
    "ES68" :'p',
    "ES69" :'p',
    "ES90" :'g',
    "ES137":'d',
    "ET86" :'p',
    "ET89" :'g',
    "ET90" :'o',
    "EU70" :'g',
    "EU71" :'p',
    "EU86" :'p',
    "EU87" :'p',
    "EU89" :'o',
    "EU90" :'o',
    "EU91" :'p',
    "EU92" :'p',
    "EU133":'p',
    "EU134":'p',
    "EV92" :'p',
    "EV85" :'p',
    "EV86" :'p',
    "EV90" :'p',
    "EV91" :'p',
    "EY7"  :'g',
    "EY74" :'d',
    "EY79" :'y',
    "EZ74" :'d',

    "FB91" :'l',
    "FC92" :'l',
    "FC95" :'l',
    "FC96" :'l',
    "FD66" :'d',
    "FD84" :'y',
    "FD91" :'l',
    "FE65" :'b',
    "FE67" :'d',
    "FE70" :'d',
    "FE72" :'m',
    "FE73" :'m',
    "FE74" :'d',
    "FF65" :'b',
    "FF70" :'d',
    "FF71" :'m',
    "FF72" :'m',
    "FF95" :'l',
    "FF104":'l',
    "FG69" :'d',
    "FG70" :'w',
    "FG71" :'m',
    "FG73" :'b',
    "FG78" :'d',
    "FG94" :'l',
    "FG96" :'l',
    "FG104":'l',
    "FG144":'d',
    "FG145":'d',
    "FH65" :'b',
    "FH70" :'d',
    "FH72" :'b',
    "FI5"  :'b',
    "FI12" :'o',
    "FI65" :'b',
    "FI74" :'b',
    "FJ66" :'d',
    "FJ75" :'w',
    "FJ99" :'w',
    "FQ66" :'m',
    "FQ68" :'m',
    "FQ70" :'m',
    "FR69" :'m',
    "FR72" :'m',
    "FS7"  :'g',
    "FX33" :'p',
    "FZ43" :'l',

    "GB32" :'p',
    "GD68" :'m',
    "GE67" :'m',
    "GE70" :'m',
    "GE72" :'m',
    "GE74" :'m',
    "GE77" :'m',
    "GH7"  :'t',
    "GK65" :'g',
    "DH149":'p',
    "FE145":'m',
    "FF145":'d',
    "FG145":'m',
    "FF146":'m',
    "FH143":'d',
    "FI142":'m',
    "FH145":'m',
    "FI145":'w',
    "FH146":'w',
    "BU146":'m',
    "DN148":'p',
    "DA150":'y',
    "DB151":'y',
    "DC150":'y',
    "DG150":'y',
    "DF154":'y',
    "DE155":'y',
    "DF156":'y',
    }
for cell, color in corrections.items():
    writeToCell(cell, color)

def resetYellowGreenColumn():
    for row in range(33, 1 + ROWS//2):
        cell = "FM33"
        col = ws[cell].column
        if row % 3 == 0:
            color = 'y'
        else:
            color = 'g'
        rgb = codes[color]
        color = f"00{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
        fill = PatternFill(start_color=f"{color}", fill_type = "solid")
        ws.cell(row, col).fill = fill

duplicate({'start':"FL46", 'end':"GP85", 'target':"FL86"})
duplicate({'start':"FL46", 'end':"GP85", 'target':"FL126"})
duplicate({'start':"FL45", 'end':"GP83", 'target':"FL162"})
resetYellowGreenColumn()

# this next set duplicates the very top of the rug, but
# the butterflies will go wrong because they don't exactly repeat
#  This is corrected below
target = newCell("EA1", diff("EA1", "FO1"))
duplicate({'start':"EA1",  'end':"FO32", 'target':target})
target = newCell("EA1", diff("EA1", "FO1"))
duplicate({'start':"EA1",  'end':"FO32", 'target':target})
target = newCell(target, diff("EA1", "FO1"))
duplicate({'start':"EA1",  'end':"FO32", 'target':target})

# Now correct the butterfiles
# duplicate 1 extra row, so we overwrite with extra column
duplicate({'start':"FH11",  'end':"FN31", 'target':"DT11"})
# now overwrite the last row with the extra column
extraColumn = 'ooomopdlmmmmmmmdldpyo'
for row, color in enumerate(extraColumn):
    cell = f"DT{row+11}"
    writeToCell(cell, color)
# duplicate middle of butterfly, again with an extra row that will be changed
duplicate({'start':"FC11",  'end':"FH31", 'target':"DN11"})

# now change to extra column
for row, color in enumerate(extraColumn):
   cell = f"DN{row+11}"
   writeToCell(cell, color)
# finally duplicate rest of butterflies to middle of rug
duplicate({'start':"EL11",  'end':"FC31", 'target':"CV11"})

# changes to be applied after butterflies reconstructed
additionalCorrections = {
    "BT17": 'd',
    "CN26": 'p',
    "DC26": 'p',
    "DO14": 'o',
    "DS14": 'o',
    "DW17": 'd',
    "EA32": 'b',
    "EG26": 'p',
    "ES26": 'p',
    "FK17": 'd',
}

for cell, color in additionalCorrections.items():
    writeToCell(cell, color)

# quadrant 2 is now complete, so duplicate quadrants 1, 3 and 4
duplicateQuadrants()

FILENAME="MASTER"
wb.save(f'data/{FILENAME}.xlsx')
print("duplication and corrections completed")
