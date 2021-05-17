from openpyxl import Workbook
wb = Workbook()

ws1 = wb.active  # get the first (and only) worksheet
ws2 = wb.create_sheet("Sheet C")     # insert at the end (default)
ws0 = wb.create_sheet("Sheet A", 0)  # insert at first position
ws1.title = "Sheet B"
print(wb.get_sheet_names())
wb.save('data/empty.xlsx')