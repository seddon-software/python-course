'''
In this exercise you create a calendar and write it to a spreadsheet.
You can achieve this in a number of ways, but we would like you to use iterators to complete this exercise.

Define the following iterators:
    monthIterator:  yields successive month names:        January, February, March ...
    cellIterator:   yields successive spreadsheet cells:  A1,B1,C1,D1,E1,F1,G1,A2,B2 ...
    
You should use the calendar module to get the month and day names and use:
        yeardays = calendar.Calendar().yeardayscalendar(year)

to help with formatting. 

Take a look at the file "calendar.xlsx" to see the expected output.
''' 

import calendar, os
import openpyxl
import itertools
YEAR = 2026

def monthIterator():
    months = list(calendar.month_name) 
    months = months[1:]     # skip blank in first entry
    for month in months:
        yield month

def cellIterator():
    for number in itertools.count():
        for letter in "ABCDEFG": 
            yield f"{letter}{number}"

def main():
    def openWorkbook():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{YEAR}"
        return wb,ws

    def saveWorkbook(wb):
        wb.save(fileName)
        cmd = f"libreoffice {fileName}"
        os.system(cmd)
    def skip(n):
        for _ in range(n): next(theCell)

    # start
    fileName = 'calendar.xlsx'
    workbook, worksheet = openWorkbook()

    # create iterators
    theCell = cellIterator()
    theMonth = monthIterator()

    days = " ".join(list(calendar.day_abbr))

    yeardays = calendar.Calendar().yeardayscalendar(YEAR)
    for months in yeardays:
        for month in months:
            skip(7)
            worksheet[next(theCell)] = next(theMonth)
            skip(6)
            for theDay in calendar.day_abbr:
                worksheet[next(theCell)] = theDay

            for week in month:
                for day in week:
                    if day == 0: day = ' '
                    worksheet[next(theCell)] = f"{day}"

    saveWorkbook(workbook)

main()

