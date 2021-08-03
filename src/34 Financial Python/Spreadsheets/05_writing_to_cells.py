from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime

wb = Workbook()
ws = wb.active
ws.title = "writing to cells"

ws['A1'] = datetime.datetime.now()
# ws['A1'].number_format = 'yyyy-mm-dd h:mm:ss'
ws['A1'].number_format = 'dd-mm-yyyy'
ws['B1'] = 42
c = ws.cell(row=3, column=1)
c.value = 43
c = ws.cell(row=1, column=4)
c.value = 44

# ws.append(iterable)
#    if iterable is a list: all values are added, starting from the first column
#    if iterable is a dict: values are assigned to the columns indicated by the keys

# append a list 5 times (rows 1-5)
for x in range(5):    
    ws.append(list(range(25, 60, 3)))
# append a dict 5 times (rows 6-10)
for x in range(5):    
    ws.append({ x:x**2 for x in range(2, 20, 3)}) # appending in cols 2, 5, 8, 11, 14, 17
 
# now add items at specific rows and cols
for row in range(20, 30):
    for col in range(1, 10):
        _ = ws.cell(column=col, row=row, value="--{}--".format(get_column_letter(col)))

wb.save('data/writingToCells.xlsx')