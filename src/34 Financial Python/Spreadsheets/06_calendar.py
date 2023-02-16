from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import datetime

daysOfTheWeek = ['M','Tu','W','Th','F','Sa','Su']
cellsNames = "ABCDEFGHIJKLMNOP"
daysInMonth = [31,28,31,30,31,30,31,31,30,31,30,31]
monthNames = ["JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE",
          "JULY","AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"]
YEAR = 2023
wb = Workbook()
ws = wb.active
ws.title = f"{YEAR}"
    
RED = colors.COLOR_INDEX[2]
monthStyle = Font(name="Arial", size = "14", color=RED, bold=True)
dayStyle = Font(name="Arial", size = "10", color=colors.BLACK)

for col in "BCDEFGHI":
    ws.column_dimensions[col].width = 3.0

for col in "JKLMNOP":
    ws.column_dimensions[col].width = 15.0

for col, value in zip(range(2, 9), daysOfTheWeek):
    ws.cell(row=1, column=col).value = value

for col, value in zip(range(10, 17), daysOfTheWeek):
    ws.cell(row=1, column=col).value = value
    ws.cell(row=1, column=col).font = monthStyle
    ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

row = 3
for month in range(12):
    ws.cell(row=row, column=4).value = monthNames[month]
    ws.cell(row=row, column=4).font = monthStyle
    row += 1
    col = datetime.datetime(YEAR, month+1, 1).weekday() + 2
    for dayNo in range(1, daysInMonth[month]+1):
        ws.cell(row=row, column=col).value = dayNo
        ws.cell(row=row, column=col).font = dayStyle
        col += 1
        if col == 9:
            col = 2
            row += 1
    row += 2
wb.save(f'data/calendar-{YEAR}.xlsx')
        
    