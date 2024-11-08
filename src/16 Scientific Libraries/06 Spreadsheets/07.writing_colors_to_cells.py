'''
Writing Colors to Cells
=======================
Setting colours for cells is based around the "PaternFill" class:
            fill = PatternFill(start_color=f"{color}", fill_type = "solid")
            ws.cell(row, col).fill = fill

where color is defined in hex.
'''

import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl

def writeToCell(row, col, color):
    cell = f"{row}{col}"
    fill = PatternFill(start_color=f"{color}", fill_type = "solid")
    ws.cell(row, col).fill = fill

fileName = 'data/colored.xlsx'
wb = load_workbook(filename=fileName)
ws = wb.active
ws.title = "writing colors to cells"

# set sizes of cells
for col in range(1, 1000):
    letter = openpyxl.cell.cell.Cell(ws, row=3, column=col).column_letter
    ws.column_dimensions[letter].width = 4  # approx 4*7 pixels
for row in range(1, 1000):
    ws.row_dimensions[row].height = 21.0

# now color cells    
for row in range(1, 100):
    for col in range(1, 90):
        h = f"{(0xFF0000 + row * col) % 0x1000000:06X}"
        writeToCell(row, col, h)

wb.save(fileName)

os.system(f"libreoffice {fileName}")
