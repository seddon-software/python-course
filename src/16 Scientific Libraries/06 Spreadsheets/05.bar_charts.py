'''
Bar Charts
==========
Openpyxl makes it relatively easy to create bar charts.  First you have to write data to the spreadsheet and
then refer to it in the chart:
            xData = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
            yData = Reference(ws, min_col=1, min_row=2, max_row=7)

The only problem I've had is the axis titles don't get aligned properly (and sometimes they don't appear at all).
I think this is because openpyxl was tested on Excel and charts behave slightly differently in LibreOffice.
'''

import os
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
)

wb = Workbook()
ws = wb.active

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
wb = Workbook(write_only=True)
ws = wb.create_sheet()
rows = [
        ('Number', 'Batch 1', 'Batch 2'),
        (2, 10, 30),
        (3, 40, 60),
        (4, 50, 70),
        (5, 20, 10),
        (6, 10, 40),
        (7, 50, 30),
    ]
for row in rows:
    ws.append(row)
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Bar Chart"
chart.y_axis.title = 'Test number'
chart.x_axis.title = 'Sample length (mm)'
xData = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
yData = Reference(ws, min_col=1, min_row=2, max_row=7)
chart.add_data(xData, titles_from_data=True)
chart.set_categories(yData)
chart.shape = 4
ws.add_chart(chart, "B10")

fileName = 'data/barCharts.xlsx' 
wb.save(fileName)

cmd = f"libreoffice {fileName}"
os.system(cmd)