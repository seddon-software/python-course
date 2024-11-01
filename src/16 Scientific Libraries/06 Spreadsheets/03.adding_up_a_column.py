import os; os.system("clear")
'''
Adding up a Column
==================

Typically you might want to add up the values in a column.  Clearly you can add a formula in the spreadsheet to
do this, but equally you can compute the sum in Python and write the answer to the spreadsheet; this is what we
do in this example.

The first part of the example writes sample data the the spreadsheet and then closes it.  You might like to put 
a breakpoint after this part is complete (see the comment below) so you can inspect the spreadsheet.

We then wait 5 seconds before the second part of the example computes the sum and writes the answer to the 
spreadsheet.  Since the spreadsheet will still be open you will need to use File/Reload to see the changes.

Note: You can use Ctrl-Q to quit from LibreOffice
'''

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime

from openpyxl.reader.excel import load_workbook


wb = Workbook()
ws = wb.active
ws.title = "summing a column"

# add some sample data to column 1
print("adding some sample data to column 1")
for row in range(3, 20):
        c = ws.cell(row=row, column=1)
        c.value = row * 10
fileName = 'data/computing.xlsx'
wb.save(fileName)

cmd = f"libreoffice {fileName}&"
os.system(cmd)

# wait for spreadsheet to be ready
import time
time.sleep(5)

# compute sum of column 1

# put a breakpoint after this line
print("Computing sum of column 1")
wb = load_workbook(fileName, read_only=False)
ws = wb.active

sum = 0
for row in range(3, 20):
        c = ws.cell(row=row, column=1)
        sum += int(c.value)

ws['A22'] = sum
wb.save(fileName)

cmd = f"libreoffice {fileName}"
os.system(cmd)

