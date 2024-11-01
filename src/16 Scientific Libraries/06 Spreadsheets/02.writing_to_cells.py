'''
Writing to Cells
================

There are several different ways to write to cells as shown below.  The "append()" method allows you to write
to multiple cells in one go.

When appending a "list" the values get inserted in the row immediately after the last write and value are
written to succesive cells.

When appending a "dict" the values also get inserted in the row immediately after the last write.  However,
the column of the cell is determined by the corresponding key in the dict (which must be an int).

Note: You can use Ctrl-Q to quit from LibreOffice.
'''

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime

fileName = 'data/writingToCells.xlsx'
wb = Workbook()
ws = wb.active
ws.title = "writing to cells"

# various ways of writing to cells
ws['A1'] = datetime.datetime.now()
ws['A1'].number_format = 'yyyy-mm-dd h:mm:ss'
ws['B1'] = 42
c = ws.cell(row=3, column=1)
c.value = 43
c = ws.cell(row=1, column=4)
c.value = 44

# append a list 5 times (rows 1-5)
#    if iterable is a list: all values are inserted, starting from the first column
for x in range(5):    
    ws.append(list(range(25, 60, 3)))
# append a dict 5 times (rows 6-10)
#    if iterable is a dict: values are assigned in the next available row, and to the columns indicated by the keys
for n in range(5):    
    ws.append({ col:f"{col}**2 = {col**2}" for col in range(2, 20, 3)}) # appending in cols 2, 5, 8, 11, 14, 17
 
# now add items at specific rows and cols
for row in range(20, 30):
    for col in range(1, 10):
        _ = ws.cell(column=col, row=row, value=f"cell {get_column_letter(col)}{row}")

wb.save(fileName)

cmd = f"libreoffice {fileName}"
os.system(cmd)

