'''
Charts
======
2D and 3D charts
'''

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import (
    AreaChart3D,
    AreaChart,
    Reference,
)


def addTitle(cell1, cell2, text):
    'cell1 and cell2 are names of cells, e.g C9, D11'
    cell = ws[cell1]
    cell.value = "Area Chart"
    cell.font = Font(name="Arial", size=24, bold=True)
    merge = f"{cell1}:{cell2}"
    ws.merge_cells(merge)
 
fileName = 'data/charts.xlsx'
wb = Workbook()
ws = wb.active

rows = [
    ['Category', 'Batch 1', 'Batch 2', 'Batch 3'],
    ['A', 70, 130, 225],
    ['B', 40, 125, 230],
    ['C', 80,  90, 235],
    ['D', 30, 100, 245],
    ['E', 25, 155, 290],
    ['F', 50, 110, 260],
]

for row in rows:
    ws.append(row)

chart = AreaChart()
chart.title = "Area Chart"
chart.style = 13
chart.x_axis.title = 'Test'
chart.y_axis.title = 'Percentage'

# data needs to start on row=1 to include titles
categories = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=7)
data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=7)
chart.add_data(data, titles_from_data=True)

# 2D Chart
chart.set_categories(categories)
addTitle("C9", "E10", "Area Chart")
ws.add_chart(chart, "C12")

# 3D Area Chart
chart = AreaChart3D()
categories = Reference(ws, min_col=1, min_row=1, max_row=7)
data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
addTitle("L9", "N10", "3D Area Chart")
ws.add_chart(chart, "L12")

# save chart and open spreadsheet
wb.save(fileName)
os.system(f"libreoffice {fileName}")
