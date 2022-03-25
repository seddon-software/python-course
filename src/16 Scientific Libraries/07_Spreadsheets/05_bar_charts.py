import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    AreaChart3D,
    AreaChart,
    Reference,
    Series,
)

wb = Workbook()
ws = wb.active

from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
wb = Workbook(write_only=True)
ws = wb.create_sheet()
rows = [
        ('Number', 'Batch 1', 'Batch 2'),
        (2, 10, 30),
        (3, 40, 60),
        (4, 50, 70),
        (5, 20, 10),
        (6, 10, 40),
        (7, 50, 30),
    ]
for row in rows:
    ws.append(row)
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Bar Chart"
chart.y_axis.title = 'Test number'
chart.x_axis.title = 'Sample length (mm)'
xData = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
yData = Reference(ws, min_col=1, min_row=2, max_row=7)
chart.add_data(xData, titles_from_data=True)
chart.set_categories(yData)
chart.shape = 4
ws.add_chart(chart, "B10")

fileName = 'data/barCharts.xlsx' 
wb.save(fileName)

cmd = f"libreoffice {fileName}"
os.system(cmd)