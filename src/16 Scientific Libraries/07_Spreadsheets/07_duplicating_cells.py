import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl

fileName = 'data/colored.xlsx'
wb = load_workbook(filename=fileName)
ws = wb.active
ws.title = "writing to cells"

# width = 1 => 7 pixels


for col in range(1, 1000):
    letter = openpyxl.cell.cell.Cell(ws, row=3, column=col).column_letter
    ws.column_dimensions[letter].width = 4
for row in range(1, 1000):
    ws.row_dimensions[row].height = 21.0
    

# ws.column_dimensions['A':'Z'].width = 4.0     # 0.98 cm
# ws.row_dimensions[1:8].height = 21.6        # 1.06 cm
ws['A2'].fill = PatternFill(start_color="FF0000", fill_type = "solid")
ws['B1'].fill = PatternFill(start_color="00FF00", fill_type = "solid")
ws['C2'].fill = PatternFill(start_color="0000FF", fill_type = "solid")
wb.save(fileName)

cmd = f"libreoffice {fileName}"
os.system(cmd)
