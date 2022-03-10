from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl

def writeToCell(row, col, color):
    cell = f"{row}{col}"
    row = ws[cell].row
    col = ws[cell].column
    fill = PatternFill(start_color=f"{color}", fill_type = "solid")
    ws.cell(row, col).fill = fill

wb = load_workbook(filename='data/colored.xlsx')
ws = wb.active
ws.title = "writing colors to cells"

# set sizes of cells
for col in range(1, 1000):
    letter = openpyxl.cell.cell.Cell(ws, row=3, column=col).column_letter
    ws.column_dimensions[letter].width = 4  # approx 4*7 pixels
for row in range(1, 1000):
    ws.row_dimensions[row].height = 21.0

# now color cells    
rows = [c for c in "ABCDEFGHIJKLM"]
for row in rows:
    for col in range(1, 30):
        writeToCell(row, col, "FF0000")
    for col in range(30, 60):
        writeToCell(row, col, "00FF00")
    for col in range(60, 90):
        writeToCell(row, col, "0000FF")

wb.save('data/colored.xlsx')
