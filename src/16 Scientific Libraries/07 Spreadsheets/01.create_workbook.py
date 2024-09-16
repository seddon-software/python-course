'''
Create Workbook
===============

To create a workbook use:
            wb = Workbook()

This will create an empty workbook with one sheet (the active sheet).  The "create_sheet()" method is used to add
further sheets.
'''

import os
from openpyxl import Workbook
wb = Workbook()

fileName = 'data/empty.xlsx'
ws1 = wb.active  # get the first (and only) worksheet
ws2 = wb.create_sheet("Sheet A3")     # insert at the end (default)
ws0 = wb.create_sheet("Sheet A1", 0)  # insert at first position
ws1.title = "Sheet A2"
print(wb.get_sheet_names())
wb.save(fileName)

cmd = f"libreoffice {fileName}"
os.system(cmd)
