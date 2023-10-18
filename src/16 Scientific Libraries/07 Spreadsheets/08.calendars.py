'''
Writing to Cells
================

Still working on this

Note: You can use Ctrl-Q to quit from LibreOffice.
'''

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime
import calendar

fileName = 'data/writingToCells.xlsx'
wb = Workbook()
ws = wb.active
ws.title = "writing to cells"

# fill in cells rows 1..60 and columns 1..7 with dummy data 
for x in range(60):    
    ws.append(list(range(8)))

allcells = []
for first, *others in ws.iter_rows():
    cells = []
    cells.append(first)
    for other in others:
        cells.append(other)
    allcells.append(cells)

# build a list of all the days in the year
days = []
c = calendar.Calendar()
for month in range(1, 13):
    g = c.itermonthdays(2023, month)
    try:
        while g:    
            days.append(next(g))
    except:
        pass

# set the correct day in all the cells
for rowcell in allcells:
    for cell in rowcell:
        try:
            value = days.pop(0)
            if value == 0: value = ""
            cell.value = value
        except:
            pass

wb.save(fileName)
cmd = f"libreoffice {fileName}"
os.system(cmd)

