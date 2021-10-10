from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter,column_index_from_string
import datetime
import calendar

wb = Workbook()
ws = wb.active
year = 2021
ws.title = f"{year}"


def nextDate(previous):
    # move on by one day
    return datetime.date.fromordinal(previous.toordinal() + 1)

# set font and colors for cells
RED = colors.COLOR_INDEX[2]
monthStyle = Font(name="Arial", size="14", color=RED, bold=True)
dayStyle = Font(name="Arial", size="10", color=colors.BLACK)

# date range
dmy = datetime.date(year=year, month=1, day=1)
dmy_end = datetime.date(year=year+1, month=1, day=1)

# set up column widths
for col in range(7):
    letter = get_column_letter(col+1)
    ws.column_dimensions[letter].width = 3.0
    letter = get_column_letter(col+8)    
    ws.column_dimensions[letter].width = 15.0

row = 1
daysOfTheWeek = ['M', 'Tu', 'W', 'Th', 'F', 'Sa', 'Su']

# set up titles for days
for col in range(7):
    ws.cell(row=row, column=col+1).value = daysOfTheWeek[col]
    ws.cell(row=row, column=col+1).font = dayStyle
    ws.cell(row=row, column=col+1).alignment = Alignment(horizontal='center')

while dmy < dmy_end:
    col = dmy.weekday()+1
    if col == 1: row += 1
    if dmy.day == 1: 
        row += 2
        monthName = calendar.month_name[dmy.month]
        ws.cell(row=row, column=3).value = monthName
        ws.cell(row=row, column=3).font = monthStyle
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        row += 1

    ws.cell(row=row, column=col).value = dmy.day
    ws.cell(row=row, column=col).font = dayStyle
    dmy = nextDate(dmy)

wb.save('data/calendar.xlsx')
