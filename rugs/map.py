#! /home/chris/anaconda3/bin/python

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.cell import Cell
import openpyxl, re

def setRowAndColumnSizes():
    for col in range(1, 1000):
        letter = Cell(ws, row=3, column=col).column_letter
        ws.column_dimensions[letter].width = 4/2
    for row in range(1, 1000):
        ws.row_dimensions[row].height = 21.0/2


def writeToSpreadsheet(x1, x2, y1, y2, value):
    for r in range(x1, x2+1):
        for c in range(y1, y2+1):
            try:
                row = r + 1
                col = c + 1
                ws.cell(row, col).value = value
                fontStyle = Font(size = "6")
                ws.cell(row, col).font = fontStyle
            except Exception as e:
                print(e)

df = pd.read_csv("map", skipinitialspace = True)
print(df)
wb = openpyxl.Workbook() 
ws = wb.active
setRowAndColumnSizes()


for index, row in df.iterrows():
    if row['fig'] != "1":
        x = int(row['x'])
        h = int(row['h'])
        y = int(row['y'])
        w = int(row['w'])
        writeToSpreadsheet(x, x+h, y, y+w, row['fig'])
wb.save(f'data/map.xlsx')

