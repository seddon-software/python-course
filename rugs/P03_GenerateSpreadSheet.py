'''
Generate Spreadsheet
'''
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import Cell
import openpyxl, re


def getBaseCoords():
    global FIG, COLS, TYPE, FLIP, X, Y
    try:
        f = open(f'data/rug{RUG}.norm', "r")
        # pick up parameters
        matcher = re.compile(r"^FIG=(.*)$").search(f.readline())
        FIG = matcher.group(1)
        matcher = re.compile(r"^COLS=(.*)$").search(f.readline())
        COLS = int(matcher.group(1))
        matcher = re.compile(r"^TYPE=(.*)$").search(f.readline())
        TYPE = matcher.group(1)
        matcher = re.compile(r"^FLIP=(.*)$").search(f.readline())
        FLIP = matcher.group(1)
        matcher = re.compile(r"^X=(.*)$").search(f.readline())
        X = matcher.group(1)
        matcher = re.compile(r"^Y=(.*)$").search(f.readline())
        Y = matcher.group(1)
        return X, Y
    except IOError as e:
        print(e)
    finally:
        f.close()

def copyNumpyArrayToSpreadsheet(baseRow, baseCol):
    fig = np.load(f"data/fig{RUG}.npy")
    print(f"data/fig{RUG}: shape of Numpy array: {fig.shape}, X={X},Y={Y}")
    for r in range(fig.shape[0]):
        for c in range(fig.shape[1]):
            red = fig[r, c, 0]
            green = fig[r, c, 1]
            blue = fig[r, c, 2]
            row = int(baseRow) + r + 1
            col = int(baseCol) + c + 1
            value = f"{red:02X}{green:02X}{blue:02X}"
            z = Cell(ws, row=row, column=col).column_letter
            cellId = f"{z}{row}"
            ws[cellId].fill = PatternFill(start_color=f"{value}", fill_type = "solid")

def getActiveWorkbookAndSheet():
    try:
        wb = load_workbook(filename=f'data/{FILENAME}.xlsx')
    except:
        wb = openpyxl.Workbook() 

    return wb, wb.active

def setRowAndColumnSizes():
    SCALE_FACTOR = 0.5
    for col in range(1, 1000):
        letter = Cell(ws, row=1, column=col).column_letter
        ws.column_dimensions[letter].width = 4*SCALE_FACTOR
    for row in range(1, 1000):
        ws.row_dimensions[row].height = 21.0*SCALE_FACTOR


def processData():
    baseRow, baseCol = getBaseCoords()
    copyNumpyArrayToSpreadsheet(baseRow, baseCol)
    wb.save(f'data/{FILENAME}.xlsx')

FILENAME="MASTER"
wb, ws = getActiveWorkbookAndSheet()
setRowAndColumnSizes()

for RUG in (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11):
     processData()
